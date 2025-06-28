import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
from datetime import datetime, timedelta
import io
from typing import List, Dict, Any

# Page configuration
st.set_page_config(
    page_title="スケジュール変換ツール",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 3rem;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        margin-bottom: 2rem;
    }
    .step-container {
        border: 2px solid #667eea;
        border-radius: 10px;
        padding: 1rem;
        margin: 1rem 0;
        background-color: #f8f9fa;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .error-box {
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        border-radius: 5px;
        padding: 1rem 0;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

class ScheduleConverter:
    def __init__(self):
        self.parsed_schedule = []
        
    def parse_schedule_text(self, text: str) -> List[Dict[str, Any]]:
        """Parse Japanese schedule text format with enhanced handling for missing time/location"""
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        schedule = []
        
        # Extract year and month from date range
        year_month_match = re.search(r'(\d{4})年(\d{1,2})月', lines[0])
        year = int(year_month_match.group(1)) if year_month_match else datetime.now().year
        month = int(year_month_match.group(2)) if year_month_match else datetime.now().month
        
        current_date = None
        current_full_date = None
        i = 0
        
        # Find where the actual schedule data starts
        while i < len(lines):
            line = lines[i]
            date_match = re.match(r'^(\d{1,2})\([日月火水木金土]\)$', line)
            if date_match:
                break
            i += 1
        
        while i < len(lines):
            line = lines[i]
            
            # Check for date line (format: "23(月)")
            date_match = re.match(r'^(\d{1,2})\([日月火水木金土]\)$', line)
            if date_match:
                current_date = line
                day_num = int(date_match.group(1))
                current_full_date = datetime(year, month, day_num)
                i += 1
                continue
            
            # Check for content in parentheses (no time/location) - handle both Japanese and English parentheses
            # Patterns: "（梱包資材購入・試作機器購入）" or "(梱包資材購入・試作機器購入)"
            parentheses_match = re.match(r'^[（(](.+)[）)]
            
            # Check for time and activity line (format: "08:50 川口本部")
            time_activity_match = re.match(r'^(\d{1,2}):(\d{2})\s+(.+)$', line)
            if time_activity_match and current_full_date:
                time = f"{time_activity_match.group(1)}:{time_activity_match.group(2)}"
                activity_text = time_activity_match.group(3).strip()
                
                # Parse location and activity
                location = ''
                activity = ''
                
                if '（' in activity_text and '）' in activity_text:
                    # Text in Japanese parentheses is usually activity - remove parentheses
                    activity = re.sub(r'[（）()]', '', activity_text)
                    location = ''
                elif '(' in activity_text and ')' in activity_text:
                    # Text in English parentheses is usually activity - remove parentheses
                    activity = re.sub(r'[（）()]', '', activity_text)
                    location = ''
                elif activity_text == '社用車帰宅':
                    location = '社用車帰宅'
                    activity = ''
                else:
                    # Split by Japanese space or multiple spaces
                    parts = re.split(r'[　\s]+', activity_text)
                    location = parts[0] if parts else ''
                    activity = ' '.join(parts[1:]) if len(parts) > 1 else ''
                
                # Handle special case where there's no activity but just location
                if not activity and location:
                    if any(keyword in location for keyword in ['打合せ', '会議', '見学', '参加', '食事', '手配', '対応']):
                        activity = location
                        location = ''
                
                # Determine if this entry has all data (time, location, activity)
                has_all_data = bool(time and (location or activity))
                
                schedule.append({
                    'date': current_date,
                    'full_date': current_full_date,
                    'time': time,
                    'location': location,
                    'activity': activity,
                    'has_all_data': has_all_data
                })
            
            i += 1
        
        return schedule
    
    def generate_filename(self, text: str) -> str:
        """Generate filename from schedule date range"""
        date_range_match = re.search(
            r'(\d{4})年(\d{1,2})月(\d{1,2})日\([日月火水木金土]\)\s*～\s*(\d{4})年(\d{1,2})月(\d{1,2})日\([日月火水木金土]\)',
            text
        )
        
        if date_range_match:
            start_year = date_range_match.group(1)
            start_month = date_range_match.group(2).zfill(2)
            start_day = date_range_match.group(3).zfill(2)
            end_year = date_range_match.group(4)
            end_month = date_range_match.group(5).zfill(2)
            end_day = date_range_match.group(6).zfill(2)
            
            return f"{start_year}{start_month}{start_day}to{end_year}{end_month}{end_day}.xlsx"
        
        # Fallback
        return f"weekly_schedule_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    def create_excel_file(self, template_file, schedule_data: List[Dict]) -> bytes:
        """Create Excel file with schedule data - prioritizing complete entries"""
        # Load template
        wb = load_workbook(template_file)
        ws = wb.active
        
        # Get date range from schedule data
        if not schedule_data:
            return wb
        
        # Extract start date from first entry
        first_entry = min(schedule_data, key=lambda x: x['full_date'])
        start_date = first_entry['full_date']
        
        # Generate all 7 days of the week starting from start_date
        all_dates = []
        for i in range(7):
            date = start_date + timedelta(days=i)
            all_dates.append(date)
        
        # Group schedule by date
        schedule_by_date = {}
        for item in schedule_data:
            date_key = item['full_date'].strftime('%Y-%m-%d')
            if date_key not in schedule_by_date:
                schedule_by_date[date_key] = []
            schedule_by_date[date_key].append(item)
        
        # Sort entries within each date: complete entries first, then incomplete ones
        for date_key in schedule_by_date:
            schedule_by_date[date_key].sort(key=lambda x: (not x.get('has_all_data', True), x['time'] or '99:99'))
        
        # Japanese day names
        day_names = ['月', '火', '水', '木', '金', '土', '日']
        
        # Populate Excel sheet with fixed 6-row spacing per day
        for day_idx, date in enumerate(all_dates):
            date_key = date.strftime('%Y-%m-%d')
            day_entries = schedule_by_date.get(date_key, [])  # Empty list if no entries
            
            # Calculate starting row for this day (6-row blocks: 7-12, 13-18, 19-24, etc.)
            day_start_row = 7 + (day_idx * 6)  # 7, 13, 19, 25, 31, 37, 43...
            
            # Get day of week (0=Monday, 6=Sunday)
            weekday_idx = date.weekday()
            day_name = day_names[weekday_idx]
            
            # Format the date column for each row in the 6-row block
            for row_offset in range(6):
                current_row = day_start_row + row_offset
                
                if row_offset == 0:
                    # First row: Day name in parentheses (月)
                    ws.cell(row=current_row, column=1, value=f"({day_name})")
                elif row_offset == 1:
                    # Second row: Full date YYYY/MM/DD
                    ws.cell(row=current_row, column=1, value=date.strftime('%Y/%m/%d'))
                else:
                    # Rows 3-6: Leave blank
                    ws.cell(row=current_row, column=1, value="")
            
            # Process entries for this day (already sorted with complete entries first)
            row_index = day_start_row
            for entry in day_entries:
                # Set time (Column B) - leave blank if no time
                if entry['time']:
                    time_parts = entry['time'].split(':')
                    time_value = datetime(1899, 12, 30, int(time_parts[0]), int(time_parts[1]))
                    ws.cell(row=row_index, column=2, value=time_value)
                else:
                    ws.cell(row=row_index, column=2, value="")
                
                # Set location (Column C) - remove parentheses if present
                location_clean = re.sub(r'[（）()]', '', entry['location']) if entry['location'] else ''
                ws.cell(row=row_index, column=3, value=location_clean)
                
                # Set activity (Column D) - remove parentheses if present
                activity_clean = re.sub(r'[（）()]', '', entry['activity']) if entry['activity'] else ''
                ws.cell(row=row_index, column=4, value=activity_clean)
                
                row_index += 1
                
                # Don't exceed the 6-row block for this day
                if row_index >= day_start_row + 6:
                    break
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

def main():
    st.markdown('<h1 class="main-header">📅 スケジュール変換ツール</h1>', unsafe_allow_html=True)
    
    # Initialize converter
    converter = ScheduleConverter()
    
    # Sidebar for instructions
    with st.sidebar:
        st.markdown("## 📋 使用方法")
        st.markdown("""
        1. **Excelテンプレートをアップロード** - 空白のExcelファイル
        2. **スケジュールテキストを貼り付け** - PDFからコピー
        3. **解析とプレビュー** - データを確認
        4. **Excelファイルをダウンロード** - 完成したファイルを取得
        """)
        
        st.markdown("## 🔧 機能")
        st.markdown("""
        - ✅ 日本語テキストの解析
        - ✅ Excelフォーマットの保持  
        - ✅ 適切な日付列の書式設定
        - ✅ 1日6行のブロック構造
        - ✅ 日本語の日付・時間対応
        - ✅ **時間・場所欠落データの処理**
        - ✅ **完全データの優先表示**
        """)
        
        st.markdown("## 📋 対応する入力形式")
        st.markdown("""
        **完全な予定:**
        `08:50 川口本部` (時間 + 場所)
        
        **活動のみ:**
        `（梱包資材購入・試作機器購入）` または `(梱包資材購入・試作機器購入)` (時間・場所なし)
        
        **括弧内容:**
        `15:00 （VE会議）` または `15:00 (VE会議)` (時間 + 活動)
        """)
        
        st.markdown("## 📋 日付列の書式")
        st.markdown("""
        **各日は6行構成:**
        - 行1: `(月)` - 曜日を括弧内に
        - 行2: `2025/6/23` - 完全な日付
        - 行3-6: スケジュール項目 (完全データを優先)
        """)
    
    # Main content area
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown('<div class="step-container">', unsafe_allow_html=True)
        st.markdown("### 📄 ステップ1: Excelテンプレートのアップロード")
        
        excel_template = st.file_uploader(
            "空白のExcelテンプレートファイル (.xlsx) を選択してください",
            type=['xlsx'],
            help="スケジュールデータを入力する空白のExcelテンプレートをアップロードしてください"
        )
        
        if excel_template:
            st.markdown('<div class="success-box">✅ Excelテンプレートが正常に読み込まれました！</div>', unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Step 2: Input schedule text
        st.markdown('<div class="step-container">', unsafe_allow_html=True)
        st.markdown("### 📱 ステップ2: スケジュールデータの入力")
        
        # Enhanced sample text with missing data cases
        sample_text = """2025年06月23日(月) ～ 2025年06月29日(日)
氏名
衣笠修平
23(月)
08:50 川口本部
15:00 （VE会議）
20:00 社用車帰宅
24(火)
10:00 海洋電子工業　打合せ　社用車
15:00 日建リース　打合せ　社用車
16:00 川口本部
25(水)
10:00 都産研台場　試作作業　久野さん同行
12:00 新橋　営業食事・打合せ
14:00 東京大丸　お中元手配
15:00 OIF多摩国分寺支所　３次元プリンタ等設備見学
26(木)
10:00 幕張メッセ　AWS展示会見学　
15:00 神保町　海洋連絡会　懇親会　参加
27(金)
（梱包資材購入・試作機器購入）
08:50 川口本部
13:00 (来客対応)
28(土)
（サンプル発送）
29(日)"""
        
        schedule_text = st.text_area(
            "スケジュールテキストをここに貼り付けてください:",
            value=sample_text,
            height=400,
            help="PDFスケジュールからテキストをコピーして貼り付けてください。時間・場所が欠落した括弧内の項目にも対応しています。"
        )
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="step-container">', unsafe_allow_html=True)
        st.markdown("### 🔍 ステップ3: 解析とプレビュー")
        
        if st.button("🔍 スケジュールを解析", type="primary"):
            if schedule_text.strip():
                try:
                    converter.parsed_schedule = converter.parse_schedule_text(schedule_text)
                    st.session_state.parsed_schedule = converter.parsed_schedule
                    st.session_state.schedule_text = schedule_text
                    
                    # Count complete vs incomplete entries
                    complete_entries = sum(1 for item in converter.parsed_schedule if item.get('has_all_data', True))
                    incomplete_entries = len(converter.parsed_schedule) - complete_entries
                    
                    st.markdown(f'<div class="success-box">✅ {len(converter.parsed_schedule)}件のスケジュール項目を正常に解析しました！<br/>📊 完全データ: {complete_entries}件 | 部分データ: {incomplete_entries}件</div>', unsafe_allow_html=True)
                except Exception as e:
                    st.markdown(f'<div class="error-box">❌ スケジュール解析エラー: {str(e)}</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="error-box">❌ 最初にスケジュールテキストを入力してください</div>', unsafe_allow_html=True)
        
        # Display parsed schedule preview
        if hasattr(st.session_state, 'parsed_schedule') and st.session_state.parsed_schedule:
            st.markdown("#### 📋 スケジュールプレビュー:")
            
            # Create preview DataFrame with priority indicator
            preview_data = []
            for item in st.session_state.parsed_schedule:
                priority = "🔴 完全" if item.get('has_all_data', True) else "🟡 部分"
                # Remove parentheses for preview display
                location_clean = re.sub(r'[（）()]', '', item['location']) if item['location'] else '-'
                activity_clean = re.sub(r'[（）()]', '', item['activity']) if item['activity'] else '-'
                
                preview_data.append({
                    '月日': item['full_date'].strftime('%m/%d (%a)'),
                    '優先度': priority,
                    'AM/PM': item['time'] or '-',
                    '訪問先': location_clean,
                    '面談内容': activity_clean
                })
            
            df = pd.DataFrame(preview_data)
            st.dataframe(df, use_container_width=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Step 4: Generate Excel file
    st.markdown('<div class="step-container">', unsafe_allow_html=True)
    st.markdown("### 📊 ステップ4: Excelファイルの生成")
    
    col3, col4 = st.columns([1, 1])
    
    with col3:
        if st.button("📊 Excelファイルを生成", type="primary", use_container_width=True):
            if not excel_template:
                st.error("最初にExcelテンプレートをアップロードしてください！")
            elif not hasattr(st.session_state, 'parsed_schedule') or not st.session_state.parsed_schedule:
                st.error("最初にスケジュールを解析してください！")
            else:
                try:
                    # Generate Excel file
                    excel_data = converter.create_excel_file(excel_template, st.session_state.parsed_schedule)
                    
                    # Generate filename
                    filename = converter.generate_filename(st.session_state.schedule_text)
                    
                    # Provide download button
                    st.download_button(
                        label="💾 Excelファイルをダウンロード",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    st.success(f"✅ Excelファイル '{filename}' が正常に生成されました！")
                    
                except Exception as e:
                    st.error(f"❌ Excelファイル生成エラー: {str(e)}")
    
    with col4:
        st.markdown('<div class="info-box">', unsafe_allow_html=True)
        st.markdown("**💡 強化された機能:**")
        st.markdown("""
        - **欠落データ対応**: `（内容）` 形式の項目
        - **スマート優先順位**: 完全データを先頭に配置  
        - **柔軟な解析**: 様々な形式に対応
        - **クリーンな出力**: すべての括弧を自動削除
        """)
        st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Footer
    st.markdown("---")
    st.markdown("### 🚀 インストールと設定")
    st.code("""
# 必要なパッケージをインストール
pip install streamlit pandas openpyxl

# アプリケーションを実行
streamlit run schedule_converter.py
    """)
    
    st.markdown("### 📋 対応フォーマット")
    st.markdown("""
    **完全な予定の形式:**
    ```
    23(月)
    08:50 川口本部
    15:00 （VE会議）
    ```
    
    **部分的な予定の形式（時間・場所欠落）:**
    ```
    27(金)
    （梱包資材購入・試作機器購入）
    08:50 川口本部
    ```
    
    **Excelでの結果:**
    - 完全データが各日のブロックの先頭に表示
    - 時間・場所の欠落セルは空白
    - 内容から括弧が自動削除
    """)

if __name__ == "__main__":
    main(), line)
            if parentheses_match and current_full_date:
                activity_text = parentheses_match.group(1).strip()
                
                schedule.append({
                    'date': current_date,
                    'full_date': current_full_date,
                    'time': '',  # No time
                    'location': '',  # No location
                    'activity': activity_text,
                    'has_all_data': False  # Mark as incomplete for sorting
                })
                i += 1
                continue
            
            # Check for time and activity line (format: "08:50 川口本部")
            time_activity_match = re.match(r'^(\d{1,2}):(\d{2})\s+(.+)$', line)
            if time_activity_match and current_full_date:
                time = f"{time_activity_match.group(1)}:{time_activity_match.group(2)}"
                activity_text = time_activity_match.group(3).strip()
                
                # Parse location and activity
                location = ''
                activity = ''
                
                if '（' in activity_text and '）' in activity_text:
                    # Text in parentheses is usually activity - remove parentheses
                    activity = re.sub(r'[（）()]', '', activity_text)
                    location = ''
                elif activity_text == '社用車帰宅':
                    location = '社用車帰宅'
                    activity = ''
                else:
                    # Split by Japanese space or multiple spaces
                    parts = re.split(r'[　\s]+', activity_text)
                    location = parts[0] if parts else ''
                    activity = ' '.join(parts[1:]) if len(parts) > 1 else ''
                
                # Handle special case where there's no activity but just location
                if not activity and location:
                    if any(keyword in location for keyword in ['打合せ', '会議', '見学', '参加', '食事', '手配', '対応']):
                        activity = location
                        location = ''
                
                # Determine if this entry has all data (time, location, activity)
                has_all_data = bool(time and (location or activity))
                
                schedule.append({
                    'date': current_date,
                    'full_date': current_full_date,
                    'time': time,
                    'location': location,
                    'activity': activity,
                    'has_all_data': has_all_data
                })
            
            i += 1
        
        return schedule
    
    def generate_filename(self, text: str) -> str:
        """Generate filename from schedule date range"""
        date_range_match = re.search(
            r'(\d{4})年(\d{1,2})月(\d{1,2})日\([日月火水木金土]\)\s*～\s*(\d{4})年(\d{1,2})月(\d{1,2})日\([日月火水木金土]\)',
            text
        )
        
        if date_range_match:
            start_year = date_range_match.group(1)
            start_month = date_range_match.group(2).zfill(2)
            start_day = date_range_match.group(3).zfill(2)
            end_year = date_range_match.group(4)
            end_month = date_range_match.group(5).zfill(2)
            end_day = date_range_match.group(6).zfill(2)
            
            return f"{start_year}{start_month}{start_day}to{end_year}{end_month}{end_day}.xlsx"
        
        # Fallback
        return f"weekly_schedule_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    def create_excel_file(self, template_file, schedule_data: List[Dict]) -> bytes:
        """Create Excel file with schedule data - prioritizing complete entries"""
        # Load template
        wb = load_workbook(template_file)
        ws = wb.active
        
        # Get date range from schedule data
        if not schedule_data:
            return wb
        
        # Extract start date from first entry
        first_entry = min(schedule_data, key=lambda x: x['full_date'])
        start_date = first_entry['full_date']
        
        # Generate all 7 days of the week starting from start_date
        all_dates = []
        for i in range(7):
            date = start_date + timedelta(days=i)
            all_dates.append(date)
        
        # Group schedule by date
        schedule_by_date = {}
        for item in schedule_data:
            date_key = item['full_date'].strftime('%Y-%m-%d')
            if date_key not in schedule_by_date:
                schedule_by_date[date_key] = []
            schedule_by_date[date_key].append(item)
        
        # Sort entries within each date: complete entries first, then incomplete ones
        for date_key in schedule_by_date:
            schedule_by_date[date_key].sort(key=lambda x: (not x.get('has_all_data', True), x['time'] or '99:99'))
        
        # Japanese day names
        day_names = ['月', '火', '水', '木', '金', '土', '日']
        
        # Populate Excel sheet with fixed 6-row spacing per day
        for day_idx, date in enumerate(all_dates):
            date_key = date.strftime('%Y-%m-%d')
            day_entries = schedule_by_date.get(date_key, [])  # Empty list if no entries
            
            # Calculate starting row for this day (6-row blocks: 7-12, 13-18, 19-24, etc.)
            day_start_row = 7 + (day_idx * 6)  # 7, 13, 19, 25, 31, 37, 43...
            
            # Get day of week (0=Monday, 6=Sunday)
            weekday_idx = date.weekday()
            day_name = day_names[weekday_idx]
            
            # Format the date column for each row in the 6-row block
            for row_offset in range(6):
                current_row = day_start_row + row_offset
                
                if row_offset == 0:
                    # First row: Day name in parentheses (月)
                    ws.cell(row=current_row, column=1, value=f"({day_name})")
                elif row_offset == 1:
                    # Second row: Full date YYYY/MM/DD
                    ws.cell(row=current_row, column=1, value=date.strftime('%Y/%m/%d'))
                else:
                    # Rows 3-6: Leave blank
                    ws.cell(row=current_row, column=1, value="")
            
            # Process entries for this day (already sorted with complete entries first)
            row_index = day_start_row
            for entry in day_entries:
                # Set time (Column B) - leave blank if no time
                if entry['time']:
                    time_parts = entry['time'].split(':')
                    time_value = datetime(1899, 12, 30, int(time_parts[0]), int(time_parts[1]))
                    ws.cell(row=row_index, column=2, value=time_value)
                else:
                    ws.cell(row=row_index, column=2, value="")
                
                # Set location (Column C) - remove parentheses if present
                location_clean = re.sub(r'[（）()]', '', entry['location']) if entry['location'] else ''
                ws.cell(row=row_index, column=3, value=location_clean)
                
                # Set activity (Column D) - remove parentheses if present
                activity_clean = re.sub(r'[（）()]', '', entry['activity']) if entry['activity'] else ''
                ws.cell(row=row_index, column=4, value=activity_clean)
                
                row_index += 1
                
                # Don't exceed the 6-row block for this day
                if row_index >= day_start_row + 6:
                    break
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

def main():
    st.markdown('<h1 class="main-header">📅 スケジュール変換ツール</h1>', unsafe_allow_html=True)
    
    # Initialize converter
    converter = ScheduleConverter()
    
    # Sidebar for instructions
    with st.sidebar:
        st.markdown("## 📋 使用方法")
        st.markdown("""
        1. **Excelテンプレートをアップロード** - 空白のExcelファイル
        2. **スケジュールテキストを貼り付け** - PDFからコピー
        3. **解析とプレビュー** - データを確認
        4. **Excelファイルをダウンロード** - 完成したファイルを取得
        """)
        
        st.markdown("## 🔧 機能")
        st.markdown("""
        - ✅ 日本語テキストの解析
        - ✅ Excelフォーマットの保持  
        - ✅ 適切な日付列の書式設定
        - ✅ 1日6行のブロック構造
        - ✅ 日本語の日付・時間対応
        - ✅ **時間・場所欠落データの処理**
        - ✅ **完全データの優先表示**
        """)
        
        st.markdown("## 📋 対応する入力形式")
        st.markdown("""
        **完全な予定:**
        `08:50 川口本部` (時間 + 場所)
        
        **活動のみ:**
        `（梱包資材購入・試作機器購入）` (時間・場所なし)
        
        **括弧内容:**
        `15:00 （VE会議）` (時間 + 活動)
        """)
        
        st.markdown("## 📋 日付列の書式")
        st.markdown("""
        **各日は6行構成:**
        - 行1: `(月)` - 曜日を括弧内に
        - 行2: `2025/6/23` - 完全な日付
        - 行3-6: スケジュール項目 (完全データを優先)
        """)
    
    # Main content area
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown('<div class="step-container">', unsafe_allow_html=True)
        st.markdown("### 📄 ステップ1: Excelテンプレートのアップロード")
        
        excel_template = st.file_uploader(
            "空白のExcelテンプレートファイル (.xlsx) を選択してください",
            type=['xlsx'],
            help="スケジュールデータを入力する空白のExcelテンプレートをアップロードしてください"
        )
        
        if excel_template:
            st.markdown('<div class="success-box">✅ Excelテンプレートが正常に読み込まれました！</div>', unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Step 2: Input schedule text
        st.markdown('<div class="step-container">', unsafe_allow_html=True)
        st.markdown("### 📱 ステップ2: スケジュールデータの入力")
        
        # Enhanced sample text with missing data cases
        sample_text = """2025年06月23日(月) ～ 2025年06月29日(日)
氏名
衣笠修平
23(月)
08:50 川口本部
15:00 （VE会議）
20:00 社用車帰宅
24(火)
10:00 海洋電子工業　打合せ　社用車
15:00 日建リース　打合せ　社用車
16:00 川口本部
25(水)
10:00 都産研台場　試作作業　久野さん同行
12:00 新橋　営業食事・打合せ
14:00 東京大丸　お中元手配
15:00 OIF多摩国分寺支所　３次元プリンタ等設備見学
26(木)
10:00 幕張メッセ　AWS展示会見学　
15:00 神保町　海洋連絡会　懇親会　参加
27(金)
（梱包資材購入・試作機器購入）
08:50 川口本部
13:00 (来客対応)
28(土)
（サンプル発送）
29(日)"""
        
        schedule_text = st.text_area(
            "スケジュールテキストをここに貼り付けてください:",
            value=sample_text,
            height=400,
            help="PDFスケジュールからテキストをコピーして貼り付けてください。時間・場所が欠落した括弧内の項目にも対応しています。"
        )
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="step-container">', unsafe_allow_html=True)
        st.markdown("### 🔍 ステップ3: 解析とプレビュー")
        
        if st.button("🔍 スケジュールを解析", type="primary"):
            if schedule_text.strip():
                try:
                    converter.parsed_schedule = converter.parse_schedule_text(schedule_text)
                    st.session_state.parsed_schedule = converter.parsed_schedule
                    st.session_state.schedule_text = schedule_text
                    
                    # Count complete vs incomplete entries
                    complete_entries = sum(1 for item in converter.parsed_schedule if item.get('has_all_data', True))
                    incomplete_entries = len(converter.parsed_schedule) - complete_entries
                    
                    st.markdown(f'<div class="success-box">✅ {len(converter.parsed_schedule)}件のスケジュール項目を正常に解析しました！<br/>📊 完全データ: {complete_entries}件 | 部分データ: {incomplete_entries}件</div>', unsafe_allow_html=True)
                except Exception as e:
                    st.markdown(f'<div class="error-box">❌ スケジュール解析エラー: {str(e)}</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="error-box">❌ 最初にスケジュールテキストを入力してください</div>', unsafe_allow_html=True)
        
        # Display parsed schedule preview
        if hasattr(st.session_state, 'parsed_schedule') and st.session_state.parsed_schedule:
            st.markdown("#### 📋 スケジュールプレビュー:")
            
            # Create preview DataFrame with priority indicator
            preview_data = []
            for item in st.session_state.parsed_schedule:
                priority = "🔴 完全" if item.get('has_all_data', True) else "🟡 部分"
                # Remove parentheses for preview display
                location_clean = re.sub(r'[（）()]', '', item['location']) if item['location'] else '-'
                activity_clean = re.sub(r'[（）()]', '', item['activity']) if item['activity'] else '-'
                
                preview_data.append({
                    '月日': item['full_date'].strftime('%m/%d (%a)'),
                    '優先度': priority,
                    'AM/PM': item['time'] or '-',
                    '訪問先': location_clean,
                    '面談内容': activity_clean
                })
            
            df = pd.DataFrame(preview_data)
            st.dataframe(df, use_container_width=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Step 4: Generate Excel file
    st.markdown('<div class="step-container">', unsafe_allow_html=True)
    st.markdown("### 📊 ステップ4: Excelファイルの生成")
    
    col3, col4 = st.columns([1, 1])
    
    with col3:
        if st.button("📊 Excelファイルを生成", type="primary", use_container_width=True):
            if not excel_template:
                st.error("最初にExcelテンプレートをアップロードしてください！")
            elif not hasattr(st.session_state, 'parsed_schedule') or not st.session_state.parsed_schedule:
                st.error("最初にスケジュールを解析してください！")
            else:
                try:
                    # Generate Excel file
                    excel_data = converter.create_excel_file(excel_template, st.session_state.parsed_schedule)
                    
                    # Generate filename
                    filename = converter.generate_filename(st.session_state.schedule_text)
                    
                    # Provide download button
                    st.download_button(
                        label="💾 Excelファイルをダウンロード",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    st.success(f"✅ Excelファイル '{filename}' が正常に生成されました！")
                    
                except Exception as e:
                    st.error(f"❌ Excelファイル生成エラー: {str(e)}")
    
    with col4:
        st.markdown('<div class="info-box">', unsafe_allow_html=True)
        st.markdown("**💡 強化された機能:**")
        st.markdown("""
        - **欠落データ対応**: `（内容）` 形式の項目
        - **スマート優先順位**: 完全データを先頭に配置  
        - **柔軟な解析**: 様々な形式に対応
        - **クリーンな出力**: すべての括弧を自動削除
        """)
        st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Footer
    st.markdown("---")
    st.markdown("### 🚀 インストールと設定")
    st.code("""
# 必要なパッケージをインストール
pip install streamlit pandas openpyxl

# アプリケーションを実行
streamlit run schedule_converter.py
    """)
    
    st.markdown("### 📋 対応フォーマット")
    st.markdown("""
    **完全な予定の形式:**
    ```
    23(月)
    08:50 川口本部
    15:00 （VE会議）
    ```
    
    **部分的な予定の形式（時間・場所欠落）:**
    ```
    27(金)
    （梱包資材購入・試作機器購入）
    08:50 川口本部
    ```
    
    **Excelでの結果:**
    - 完全データが各日のブロックの先頭に表示
    - 時間・場所の欠落セルは空白
    - 内容から括弧が自動削除
    """)

if __name__ == "__main__":
    main()